from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl


IMPORT_NOTE = "Import Excel OEA: data oea 2.xlsx"
CREATED_BY_NAME = "Robby"
FIRST_SCORE_ROW = 3
LAST_SCORE_ROW = 20


def sql_string(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def sql_date(value: datetime | date) -> str:
    if isinstance(value, datetime):
        value = value.date()
    return f"date '{value.isoformat()}'"


def normalize_date(value):
    if isinstance(value, (datetime, date)):
        return value
    raise ValueError(f"Invalid played_at value: {value!r}")


def load_rows(xlsx_path: Path):
    workbook_values = openpyxl.load_workbook(xlsx_path, data_only=True)
    workbook_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)
    sheet_values = workbook_values.active
    sheet_formulas = workbook_formulas.active

    merged_ranges = sorted(
        [
            cell_range
            for cell_range in sheet_formulas.merged_cells.ranges
            if cell_range.min_row == 1 and cell_range.max_row == 1
        ],
        key=lambda cell_range: cell_range.min_col,
    )

    if not merged_ranges:
        raise ValueError("No merged date ranges found in row 1.")

    voice_names = []
    for row_index in range(FIRST_SCORE_ROW, LAST_SCORE_ROW + 1):
        voice_name = sheet_values.cell(row_index, 1).value
        if not voice_name:
            raise ValueError(f"Missing voice name in row {row_index}.")
        voice_names.append(str(voice_name).strip())

    rows = []
    game_summaries = []

    for game_key, cell_range in enumerate(merged_ranges, start=1):
        played_at = normalize_date(sheet_values.cell(1, cell_range.min_col).value)
        players = []

        for col_index in range(cell_range.min_col, cell_range.max_col + 1):
            player_name = sheet_values.cell(2, col_index).value
            if not player_name:
                raise ValueError(
                    f"Missing player name for game {game_key}, column {col_index}."
                )
            players.append((col_index, str(player_name).strip()))

        lower_players = [player_name.lower() for _, player_name in players]
        duplicate_players = [
            player_name
            for player_name, count in Counter(lower_players).items()
            if count > 1
        ]
        if duplicate_players:
            raise ValueError(
                f"Duplicate player in game {game_key} ({played_at}): {duplicate_players}"
            )

        game_summaries.append(
            {
                "game_key": game_key,
                "played_at": played_at,
                "players": [player_name for _, player_name in players],
            }
        )

        for player_order, (col_index, player_name) in enumerate(players, start=1):
            for row_index, voice_name in zip(
                range(FIRST_SCORE_ROW, LAST_SCORE_ROW + 1), voice_names
            ):
                score = sheet_values.cell(row_index, col_index).value
                formula = sheet_formulas.cell(row_index, col_index).value

                # Excel sometimes lacks cached values for formulas. Row 9 / "Media"
                # is formula-driven and can be recomputed directly from rows 3-8.
                if (
                    score is None
                    and row_index == 9
                    and isinstance(formula, str)
                    and formula.startswith("=IF")
                ):
                    upper_total = sum(
                        sheet_values.cell(source_row, col_index).value or 0
                        for source_row in range(3, 9)
                    )
                    score = 25 if upper_total >= 63 else 0

                if (
                    score is None
                    or isinstance(score, bool)
                    or not isinstance(score, (int, float))
                    or int(score) != score
                ):
                    raise ValueError(
                        "Invalid score "
                        f"game={game_key}, date={played_at}, player={player_name}, "
                        f"voice={voice_name}, row={row_index}, col={col_index}, "
                        f"value={score!r}, formula={formula!r}"
                    )

                rows.append(
                    {
                        "game_key": game_key,
                        "played_at": played_at,
                        "player_order": player_order,
                        "player_name": player_name,
                        "voice_name": voice_name,
                        "score": int(score),
                    }
                )

    return sheet_values.title, voice_names, game_summaries, rows


def build_sql(source_path: Path, sheet_name: str, game_summaries, rows) -> str:
    values = []
    for row in rows:
        values.append(
            "\t("
            f"{row['game_key']}, "
            f"{sql_date(row['played_at'])}, "
            f"{row['player_order']}, "
            f"{sql_string(row['player_name'])}, "
            f"{sql_string(row['voice_name'])}, "
            f"{row['score']}"
            ")"
        )

    expected_games = len(game_summaries)
    expected_players = sum(len(game["players"]) for game in game_summaries)
    expected_scores = len(rows)
    value_sql = ",\n".join(values)

    return f"""-- Import OEA da {source_path}
-- Generato da Codex.
-- Workbook: {source_path.name}
-- Sheet: {sheet_name}
-- Partite: {expected_games}
-- Giocatori partita: {expected_players}
-- Punteggi: {expected_scores}
--
-- Eseguire nello SQL Editor di Supabase.
-- Lo script e' conservativo:
-- - usa un game_key interno e UUID pre-generati, quindi supporta anche piu' partite nella stessa data;
-- - fallisce se un giocatore Excel non esiste una sola volta in public.oea_users;
-- - fallisce se una voce Excel non esiste una sola volta in public.oea_voices.name;
-- - fallisce se questo stesso import risulta gia' presente tramite notes = {sql_string(IMPORT_NOTE)}.

begin;

create temp table _oea_excel_import (
\tgame_key integer not null,
\tplayed_at date not null,
\tplayer_order integer not null,
\tplayer_name text not null,
\tvoice_name text not null,
\tscore integer not null
) on commit drop;

insert into _oea_excel_import (game_key, played_at, player_order, player_name, voice_name, score)
values
{value_sql};

do $$
declare
\tmissing_users text;
\tduplicate_user_names text;
\tmissing_voices text;
\tduplicate_voice_names text;
\tcreated_by_count integer;
\talready_imported_count integer;
begin
\tselect string_agg(distinct i.player_name, ', ' order by i.player_name)
\tinto missing_users
\tfrom _oea_excel_import i
\twhere not exists (
\t\tselect 1
\t\tfrom public.oea_users u
\t\twhere lower(u.name) = lower(i.player_name)
\t);

\tif missing_users is not null then
\t\traise exception 'Import OEA bloccato: utenti non trovati in oea_users: %', missing_users;
\tend if;

\tselect string_agg(name, ', ' order by name)
\tinto duplicate_user_names
\tfrom (
\t\tselect lower(u.name) as name
\t\tfrom public.oea_users u
\t\tjoin (select distinct lower(player_name) as name from _oea_excel_import) i on i.name = lower(u.name)
\t\tgroup by lower(u.name)
\t\thaving count(*) > 1
\t) duplicates;

\tif duplicate_user_names is not null then
\t\traise exception 'Import OEA bloccato: nomi utente duplicati in oea_users: %', duplicate_user_names;
\tend if;

\tselect string_agg(distinct i.voice_name, ', ' order by i.voice_name)
\tinto missing_voices
\tfrom _oea_excel_import i
\twhere not exists (
\t\tselect 1
\t\tfrom public.oea_voices v
\t\twhere lower(v.name) = lower(i.voice_name)
\t);

\tif missing_voices is not null then
\t\traise exception 'Import OEA bloccato: voci non trovate in oea_voices.name: %', missing_voices;
\tend if;

\tselect string_agg(name, ', ' order by name)
\tinto duplicate_voice_names
\tfrom (
\t\tselect lower(v.name) as name
\t\tfrom public.oea_voices v
\t\tjoin (select distinct lower(voice_name) as name from _oea_excel_import) i on i.name = lower(v.name)
\t\tgroup by lower(v.name)
\t\thaving count(*) > 1
\t) duplicates;

\tif duplicate_voice_names is not null then
\t\traise exception 'Import OEA bloccato: nomi voce duplicati in oea_voices: %', duplicate_voice_names;
\tend if;

\tselect count(*)
\tinto created_by_count
\tfrom public.oea_users u
\twhere lower(u.name) = lower({sql_string(CREATED_BY_NAME)});

\tif created_by_count <> 1 then
\t\traise exception 'Import OEA bloccato: created_by % deve esistere una sola volta in oea_users, trovati %', {sql_string(CREATED_BY_NAME)}, created_by_count;
\tend if;

\tselect count(*)
\tinto already_imported_count
\tfrom public.oea_games g
\twhere g.notes = {sql_string(IMPORT_NOTE)};

\tif already_imported_count > 0 then
\t\traise exception 'Import OEA bloccato: risultano gia'' presenti % partite con notes = %', already_imported_count, {sql_string(IMPORT_NOTE)};
\tend if;
end $$;

create temp table _oea_created_games (
\tgame_key integer primary key,
\tgame_id uuid not null default gen_random_uuid(),
\tplayed_at date not null
) on commit drop;

insert into _oea_created_games (game_key, played_at)
select distinct game_key, played_at
from _oea_excel_import
order by game_key;

insert into public.oea_games (id, played_at, on_record, created_by, notes)
select
\tcg.game_id,
\tcg.played_at,
\ttrue,
\t(select u.id from public.oea_users u where lower(u.name) = lower({sql_string(CREATED_BY_NAME)}) limit 1),
\t{sql_string(IMPORT_NOTE)}
from _oea_created_games cg
order by cg.game_key;

insert into public.oea_game_players (game_id, user_id, player_order)
select distinct
\tcg.game_id,
\tu.id,
\ti.player_order
from _oea_excel_import i
join _oea_created_games cg on cg.game_key = i.game_key
join public.oea_users u on lower(u.name) = lower(i.player_name)
order by cg.game_id, i.player_order;

insert into public.oea_scores (game_id, user_id, voice_id, score)
select
\tcg.game_id,
\tu.id,
\tv.id,
\ti.score
from _oea_excel_import i
join _oea_created_games cg on cg.game_key = i.game_key
join public.oea_users u on lower(u.name) = lower(i.player_name)
join public.oea_voices v on lower(v.name) = lower(i.voice_name)
order by i.game_key, i.player_order, v.sort_order;

do $$
declare
\tactual_games integer;
\tactual_players integer;
\tactual_scores integer;
begin
\tselect count(*) into actual_games from _oea_created_games;
\tselect count(*)
\tinto actual_players
\tfrom public.oea_game_players gp
\tjoin _oea_created_games cg on cg.game_id = gp.game_id;
\tselect count(*)
\tinto actual_scores
\tfrom public.oea_scores s
\tjoin _oea_created_games cg on cg.game_id = s.game_id;

\tif actual_games <> {expected_games} or actual_players <> {expected_players} or actual_scores <> {expected_scores} then
\t\traise exception 'Import OEA bloccato: conteggi inattesi games %, players %, scores %; attesi %, %, %',
\t\t\tactual_games, actual_players, actual_scores,
\t\t\t{expected_games}, {expected_players}, {expected_scores};
\tend if;
end $$;

select
\t(select count(*) from _oea_created_games) as games_inserted,
\t(select count(*) from public.oea_game_players gp join _oea_created_games cg on cg.game_id = gp.game_id) as game_players_inserted,
\t(select count(*) from public.oea_scores s join _oea_created_games cg on cg.game_id = s.game_id) as scores_inserted;

commit;
"""


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("output_sql_path", type=Path)
    args = parser.parse_args()

    sheet_name, voice_names, game_summaries, rows = load_rows(args.xlsx_path)
    sql = build_sql(args.xlsx_path, sheet_name, game_summaries, rows)
    args.output_sql_path.parent.mkdir(parents=True, exist_ok=True)
    args.output_sql_path.write_text(sql, encoding="utf-8", newline="\n")

    date_values = [
        game["played_at"].date().isoformat()
        if isinstance(game["played_at"], datetime)
        else game["played_at"].isoformat()
        for game in game_summaries
    ]
    duplicate_dates = {
        played_at: count
        for played_at, count in Counter(date_values).items()
        if count > 1
    }
    print(f"sheet={sheet_name}")
    print(f"games={len(game_summaries)}")
    print(f"game_players={sum(len(game['players']) for game in game_summaries)}")
    print(f"scores={len(rows)}")
    print(f"voices={','.join(voice_names)}")
    print(f"duplicate_dates={duplicate_dates}")
    print(f"output={args.output_sql_path}")


if __name__ == "__main__":
    main()
